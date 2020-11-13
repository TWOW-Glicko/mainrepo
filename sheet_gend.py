import random
import json
import math
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from openpyxl.utils import get_column_letter
from PIL import ImageFont
from datetime import date

months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
days = [31,28,31,30,31,30,31,31,30,31,30,31]
letters = [chr(ord('A')+i) for i in range(26)]

#function for cell colors (inputs a score and outputs color value)
def color_bg(rating):
	if rating < 2000:
		rating = 2000
	if rating > 8000:
		rating = 8000
	if rating < 4000:
		color = '%0.2X%0.2X%0.2X' % (int(48*(rating-2000)/2000), 48, 0)
	elif rating < 6000:
		color = '%0.2X%0.2X%0.2X' % (48, 48-int(48*(rating-4000)/2000), 0)
	else:
		color = '%0.2X%0.2X%0.2X' % (48, 0, int(48*(rating-6000)/2000))
	return color

#function for text colors (inputs a score and outputs color value)
def color_text(rating):
	if rating < 2000:
		rating = 2000
	if rating > 8000:
		rating = 8000
	if rating < 4000:
		color = '%0.2X%0.2X%0.2X' % (207+int(48*(rating-2000)/2000), 255, 207)
	elif rating < 6000:
		color = '%0.2X%0.2X%0.2X' % (255, 255-int(48*(rating-4000)/2000), 207)
	else:
		color = '%0.2X%0.2X%0.2X' % (255, 207, 207+int(48*(rating-6000)/2000))
	return color

def inttodate(txtdate):
	yr = 2000+txtdate//10000
	mth = (txtdate//100)%100
	dy = txtdate % 100
	rldate = date(yr,mth,dy)
	return rldate


def findrating(data, rdate):	
	firstdate = inttodate(data[0])
	diff = rdate - firstdate
	
	index = diff.days+1
	rd = data[index][0]
	while len(data[index]) == 1:
		index -= 1
	rm = data[index][1]
	rp = data[index][2]
	
	score = 5*(rm-2*rd)
	
	return (score,rm,rd,rp)
	
official = input('Official? (y/n): ')
CUTOFF = 1 + 4*(official == 'y') 				#Min number of rounds to be included

#Give me a day!
print('Date to generate? (YYMMDD)')
CURRENT = int(input().rstrip())	#CURRENT = date in text form

#Add new sheet to write to "Ratings"
wb = Workbook()
ws = wb.active
ws.title = "Ratings"
	
with open('resultdaily.json',encoding = 'utf-8') as f: 
    result = json.load(f)			#Dictionary full of contestant stats [Contestant][first date,[raw RD, raw carry RM, RP, raw RD]]

sresult = dict()
ranking = []

for i in range(7):
	for name in result.keys():
		if result[name][0] <= CURRENT:
			sresult[name] = {"Score":0, "RM":0, "RD":0, "RP":0} #add name to sresult
				
			(score,rm,rd,rp) = findrating(result[name],inttodate(CURRENT))
			
			sresult[name]['Score'] = score
			sresult[name]['RM'] = rm
			sresult[name]['RD'] = rd
			sresult[name]['RP'] = rp

for name, rating in sresult.items():	#Cycles through all names (rating = list [Score, RM, RD, RP])
	score = rating['Score']
	if rating['RP'] >= CUTOFF:						#Check if RP reaches the cutoff
		ranking.append([score, 5*rating['RM'], 5*rating['RD'], rating['RP'], name])	#Add new element to list with contestant data

ranking.sort(reverse=True)						#sort ranking from highest to lowest score (to get actual rankings)

#formatting stuff
#default = dark cell style (used in the header and the rounds list)
default = NamedStyle(name='default')
default.alignment = Alignment(horizontal='center', vertical='center')
default.fill = PatternFill('solid', fgColor='000000')
default.font = Font(name='Assistant', color='e6e6e6', size=13)

#default2 = lighter cell style (used in contestant names and the RM/RD/RP columns)
default2 = NamedStyle(name='default2')
default2.alignment = Alignment(horizontal='center', vertical='center')
default2.fill = PatternFill('solid', fgColor='212121')
default2.font = Font(name='Assistant', color='e6e6e6', size=13)

#set first row size
ws.row_dimensions[1].height = 23
#just setting a bunch of column sizes
sizes = [7,25,11,11,11,11,2]
for i in range(len(sizes)):
	ws.column_dimensions[letters[i]].width = sizes[i]

#cell style for whole header
for j in range(1, 8):	
	ws.cell(row=1, column=j).style = default

#Column titles
ws['B1'].value = 'Name'
ws['C1'].value = 'Score'
ws['D1'].value = 'RM'
ws['E1'].value = 'RD'
ws['F1'].value = 'RP'

#Filling it in!
for i in range(len(ranking)):							#Cycle across all contestants in order of rank, i = rank-1
	ws.row_dimensions[i+2].height = 23					#set height of row
	ws.cell(row=i+2, column=1).style = default			#set first column black
	ws.cell(row=i+2, column=1).value = i+1				#Put ranks in first column
	for j in range(2, 8):							#Across all of the columns,
		ws.cell(row=i+2, column=j).style = default2		#set style (of cell)
		ws.cell(row=i+2, column=j).number_format = '0'	#set format	(of cell)
	ws.cell(row=i+2, column=7).style = default			#set divider black
	
	name = ranking[i][4]								#Pull contestant name								
	ws.cell(row=i+2, column=2).value = name				#Put in regular data (Name, Score, RM, RD, RP)
	ws.cell(row=i+2, column=3).value = ranking[i][0]
	ws.cell(row=i+2, column=4).value = ranking[i][1]
	ws.cell(row=i+2, column=5).value = ranking[i][2]
	ws.cell(row=i+2, column=6).value = ranking[i][3]
	
	value = ws.cell(row=i+2, column=3).value

	ws.cell(row=i+2, column=3).fill = PatternFill('solid', fgColor=color_bg(value))
	ws.cell(row=i+2, column=3).font = Font(name='Assistant', color=color_text(value), size=13)
		


wb.save("dailyglicko.xlsx")