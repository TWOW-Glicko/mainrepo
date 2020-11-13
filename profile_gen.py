import random
import json
import math
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from openpyxl.utils import get_column_letter
from PIL import ImageFont

months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

def parity(num):
	numstring = '{x:>6}'.format(x = '{y:.2f}'.format(y = abs(num)))
	if num >= 0:
		numstring = '+' + numstring
	else:
		numstring = '-' + numstring
	return  numstring
	
def indexed(num):
	numstring = str(num)
	if num%10 == 1:
		numstring += 'st'
	elif num%10 == 2:
		numstring += 'nd'
	elif num%10 == 3:
		numstring += 'rd'
	else:
		numstring += 'th'
		
	return numstring
	

print('Contestant?') 				#Input Contestant
CONTST = input().rstrip() 			#CONTST = Contestant name string
print('Month?')						#Input Month (as integer)
CURRENT = int(input().rstrip())		#CURRENT = Month as int

with open('result.json',encoding = 'utf-8') as f: 
    result = json.load(f)			#Dictionary full of contestant stats [month][Contestant][raw RM, raw carry RD, RP, raw RD]
	
with open('history.json',encoding = 'utf-8') as f:
	history = json.load(f)			#Dictionary (once nested) full of round names [month][Contestant][Roundname][Scorechange,placement,#ofcontestants]
	
with open('rounds.json',encoding = 'utf-8') as f:
	rounds = json.load(f)			#Array full of dictionaries of round names with items being a 2 element array: [month][Roundname][Round strength, Date]

g = open('profiles/' + CONTST + '.txt','w',encoding = 'utf-8')	#input profile info into g

for i in range(0,CURRENT+1):			#i goes across month-1s
	if CONTST in result[i].keys():			#check if contestant has rating that month
		g.write(''+months[i%12]+" \'"+str((i+192)//12))		#month name 
		g.write('\n') 				#Spacing between MONTH and data
		
		if CONTST in history[i].keys():		#check if contestant played any rounds that month
			
			roundslist = []
			
			#Read off rounds from dictionary under Contestant
			for rnd, rating in history[i][CONTST].items():	
				roundslist.append([rnd,rating[0],rating[1],rating[2],rounds[i][rnd][1]%100])
				
			roundslist.sort(key=lambda x: x[4])
			
			#Find max round string length
			mrsl = 0
			for j in range(0,len(roundslist)):
				if len(roundslist[j][0]) > mrsl:
					mrsl = len(roundslist[j][0])
			
			#Input the rounds
			for j in range(0,len(roundslist)):		#Read off every round
				g.write('{date:<9}'.format(date = indexed(roundslist[j][4])))
				g.write('{roundname:<24}'.format(roundname = roundslist[j][0]))
				g.write('{change:<9}'.format(change = parity(round(roundslist[j][1],2))))
				g.write('{place:>4} / {contestants:<4}'.format(place = str(roundslist[j][2]), contestants = str(roundslist[j][3])))
				
				g.write('\n')
			
		g.write('\n')
		
		#Calculate changes
		changes = [0]*4
		if CONTST in result[i-1].keys():
			changes = [result[i][CONTST][j] - result[i-1][CONTST][j] for j in range(4)]
		
		#Put contestant stats
		g.write('{x:<4}'.format(x = ''))
		g.write('{x:<10}'.format(x = 'Score: ') + '{x:<15}'.format(x = round(5*(result[i][CONTST][0]-2*result[i][CONTST][3]),2)))
		g.write('Δ{x:<9}'.format(x = 'Score:') + '{x}'.format(x = parity(round(5*(changes[0]-2*changes[3]),2))))
		g.write('\n')
		
		g.write('{x:<4}'.format(x = ''))
		g.write('{x:<10}'.format(x = 'RM: ') + '{x:<15}'.format(x = round(5*result[i][CONTST][0],2)))
		g.write('Δ{x:<9}'.format(x = 'RM: ') + '{x}'.format(x = parity(round(5*changes[0],2))))
		g.write('\n')
		
		g.write('{x:<4}'.format(x = ''))
		g.write('{x:<10}'.format(x = 'RD: ') + '{x:<15}'.format(x = round(5*result[i][CONTST][3],2)))
		g.write('Δ{x:<9}'.format(x = 'RD: ') + '{x}'.format(x = parity(round(5*changes[3],2))))
		g.write('\n')
		
		g.write('{x:<4}'.format(x = ''))
		g.write('{x:<10}'.format(x = 'RP: ') + '{x:<15}'.format(x = result[i][CONTST][2]))
		g.write('Δ{x:<9}'.format(x = 'RP: ') + '{x}'.format(x = parity(changes[2])))
		g.write('\n')
		
		g.write('\n') #Additional line
		g.write('\n') 
