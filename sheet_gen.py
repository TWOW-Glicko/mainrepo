import random
import json
import math
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from openpyxl.utils import get_column_letter
from PIL import ImageFont

official = input('Official? (y/n): ')
CUTOFF = 100 + 1000*(official == "n")				#Min number of rounds to be included

months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']	#What do you think this is
letters = [chr(ord('A')+i) for i in range(26)]

def parity(num):
	if abs(num) < 0.5:
		numstring = '-'
	else:
		numstring = '{x:.0f}'.format(x = abs(num))
		if num >= 0:
			numstring = '+' + numstring
		else:
			numstring = '-' + numstring
	
	return  numstring

#function for cell colors (inputs a score and outputs color value)
def color_bg(rating, b=1):
	if rating < 2000:
		rating = 2000
	if rating > 8000:
		rating = 8000
	if rating < 4000:
		color = '%0.2X%0.2X%0.2X' % tuple(int(b*v) for v in (int(48*(rating-2000)/2000), 48, 0))
	elif rating < 6000:
		color = '%0.2X%0.2X%0.2X' % tuple(int(b*v) for v in (48, 48-int(48*(rating-4000)/2000), 0))
	else:
		color = '%0.2X%0.2X%0.2X' % tuple(int(b*v) for v in (48, 0, int(48*(rating-6000)/2000)))
	return color

#function for text colors (inputs a score and outputs color value)
def color_text(rating, b=1):
	if rating < 2000:
		rating = 2000
	if rating > 8000:
		rating = 8000
	if rating < 4000:
		color = '%0.2X%0.2X%0.2X' % tuple(int(b*v) for v in (207+int(48*(rating-2000)/2000), 255, 207))
	elif rating < 6000:
		color = '%0.2X%0.2X%0.2X' % tuple(int(b*v) for v in (255, 255-int(48*(rating-4000)/2000), 207))
	else:
		color = '%0.2X%0.2X%0.2X' % tuple(int(b*v) for v in (255, 207, 207+int(48*(rating-6000)/2000)))
	return color

#Give me a month!
print('Year?')
yr = int(input().rstrip())
print('Month?')
mth = int(input().rstrip())

CURRENT = mth + (yr-16)*12 - 1	#CURRENT = month number

#Add new sheet to write to "Ratings"
wb = Workbook()
ws = wb.active
ws.title = "Ratings"

with open('result.json',encoding = 'utf-8') as f:
    result = json.load(f)		#Dictionary full of contestant stats [month][Contestant][raw RM, raw carry RD, RP, raw RD]
	
with open('history.json',encoding = 'utf-8') as f:
	history = json.load(f)		#Dictionary (once nested) full of round names [month][Contestant][Roundname][Scorechange,placement,#ofcontestants]
	
with open('rounds.json',encoding = 'utf-8') as f:
	rounds = json.load(f)		#Dictionary (once nested) full of round strengths [month][Roundname][Strength]

shistory = dict()				#shistory is dict full of historical contestant stats [Contestant][Score, RM, RD, RP][month]
ranking = []					#ranking is list version of result dictionary for CURRENT month [rank-1][Score, RM, RD, RP, Contestant]
								#(to be put into the sheet)

for i in range(CURRENT+1):						#cycles through months
	for name, rating in result[i].items():		#cycles through all names (rating = list [Score, RM, RD, RP]
		if name not in shistory:				#add name to shistory if not yet there and fill in a CURRENT+1 size empty list
			shistory[name] = {"Score":[None]*(CURRENT+1), "RM":[None]*(CURRENT+1), "RD":[None]*(CURRENT+1), "RP":[None]*(CURRENT+1)}
		score = 5*(rating[0]-2*rating[3])		#add stats for that month
		shistory[name]['Score'][i] = score		
		shistory[name]['RM'][i] = 5*rating[0]
		shistory[name]['RD'][i] = 5*rating[3]
		shistory[name]['RP'][i] = rating[2]

for name,rating in result[CURRENT].items():
	ranking.append(5*(rating[0]-2*rating[3]))

ranking.sort(reverse = True)
top100 = ranking[100]

ranking = []
for name, rating in result[CURRENT].items():	#Cycles through all names (rating = list [Score, RM, RD, RP])
	score = 5*(rating[0]-2*rating[3])
	if score > top100 or rating[3] <= CUTOFF:
		ranking.append([score, 5*rating[0], 5*rating[3], rating[2], name])	#Add new element to list with contestant data

ranking.sort(reverse=True)						#sort ranking from highest to lowest score (to get actual rankings)

#formatting stuff
#default = dark cell style (used in the header and the rounds list)
default = NamedStyle(name='default')
default.alignment = Alignment(horizontal='center', vertical='center')
default.fill = PatternFill('solid', fgColor='000000')
default.font = Font(name='Assistant', color='e6e6e6', size=13)

#default2 = lighter cell style (used in contestant names and the RM/RD/RP columns)
default2 = NamedStyle(name='default2')
default2.alignment = Alignment(horizontal='center', vertical='center')
default2.fill = PatternFill('solid', fgColor='212121')
default2.font = Font(name='Assistant', color='e6e6e6', size=13)

#darkened = faded cell style (used in contestants with >500 RD in the top 100)
darkened = NamedStyle(name='darkened')
darkened.alignment = Alignment(horizontal='center', vertical='center')
darkened.fill = PatternFill('solid', fgColor='111111')
darkened.font = Font(name='Assistant', color='727272', size=13)

#set first row size
ws.row_dimensions[1].height = 23
#just setting a bunch of column sizes
sizes = [7,25,11,11,11,11,11,2]
for i in range(len(sizes)):
	ws.column_dimensions[letters[i]].width = sizes[i]

#cell sizes / headers for historical score areas
for j in range(9, 9+CURRENT):	#start at first column after divider and go to end
	ws.column_dimensions[get_column_letter(i)].width = 11	#set column width
	val = CURRENT+8-j										#val goes from CURRENT-1 to 0
	ws.cell(row=1, column=j).value = months[val%12]+" \'"+str((val+192)//12)	#month and date (using a little bit of modulo)

#cell style for whole header
for j in range(1, 9+CURRENT):	
	ws.cell(row=1, column=j).style = default

#Column titles
ws['B1'].value = 'Name'
ws['C1'].value = 'Score'
ws['D1'].value = 'RM'
ws['E1'].value = 'RD'
ws['F1'].value = 'RP'
ws['G1'].value = 'Best'

#Number of players skipped due to RD cutoff, for rank correction
s = 0

#Filling it in!
for i in range(len(ranking)):							#Cycle across all contestants in order of rank, i = rank-1
	if ranking[i][2] <= CUTOFF*5:
		darker_row = False
	else:
		s += 1
		darker_row = True
	
	brightness = 0.5 + 0.5 * (not darker_row)
	
	ws.row_dimensions[i+2].height = 23					#set height of row
	ws.cell(row=i+2, column=1).style = default			#set first column black

	if not darker_row:									#Put rank in first column if <=500 RD
		ws.cell(row=i+2, column=1).value = i+1-s		#Omit the rank otherwise
	
	for j in range(2, 9+CURRENT):						#Across all of the columns,
		if darker_row:
			ws.cell(row=i+2, column=j).style = darkened #set darker cell style
		else:	
			ws.cell(row=i+2, column=j).style = default2	#set normal cell style
		
		ws.cell(row=i+2, column=j).number_format = '0'	#set format	(of cell)
	ws.cell(row=i+2, column=8).style = default			#set divider black
	
	name = ranking[i][4]								#Pull contestant name								
	ws.cell(row=i+2, column=2).value = name				#Put in regular data (Name, Score, RM, RD, RP)
	ws.cell(row=i+2, column=3).value = ranking[i][0]
	ws.cell(row=i+2, column=4).value = ranking[i][1]
	ws.cell(row=i+2, column=5).value = ranking[i][2]
	ws.cell(row=i+2, column=6).value = ranking[i][3]
	ws.cell(row=i+2, column=7).value = max(x for x in shistory[name]['Score'] if x is not None)	#Search through contestant history for maximum score
	
	for j in [3, 7]:									#Use color_bg and color_text to format Score and Max
		value = ws.cell(row=i+2, column=j).value
		#print(color_text(value))
		ws.cell(row=i+2, column=j).fill = PatternFill('solid', fgColor=color_bg(value, b=brightness))
		ws.cell(row=i+2, column=j).font = Font(name='Assistant', color=color_text(value, b=brightness), size=13)
		
	for j in range(9, 9+CURRENT):						#Across history columns
		value = shistory[name]['Score'][CURRENT+8-j]	#Pull historical Score
		if value is not None:							#Check if a Score exists
			ws.cell(row=i+2, column=j).value = value	#Put score in and use color_bg and color_text to format
			ws.cell(row=i+2, column=j).fill = PatternFill('solid', fgColor=color_bg(value, b=brightness))
			ws.cell(row=i+2, column=j).font = Font(name='Assistant', color=color_text(value, b=brightness), size=13)

#Video Information
ws = wb.create_sheet("Video")

#Column sizes
sizes = [7,25] + [11 for i in range(9)] + [7] + [25,11,11]*3 + [7] + [25,11,11]*3 + [7]
for i in range(len(sizes)):
	if i < 26:
		ws.column_dimensions[letters[i]].width = sizes[i]
	else:
		ws.column_dimensions[letters[i//26-1] + letters[i%26]].width = sizes[i]

#Column names
ws['A1'].value = 'Rank'
ws['B1'].value = 'Name'
ws['C1'].value = 'Score'
ws['E1'].value = 'Score Change'
ws['F1'].value = 'RM'
ws['G1'].value = 'RM Change'
ws['H1'].value = 'RD'
ws['I1'].value = 'RD Change'
ws['J1'].value = 'RP'
ws['K1'].value = 'RP Change'

for i in range(len(ranking)):										
	for j in range(1,32):
		ws.cell(row = 1,column = j).style = default
		ws.cell(row = i+2,column = j).style = default2
	ws.cell(row = i+2,column = 1).value = i+1
	name = ranking[i][4]
	ws.cell(row=i+2, column=2).value = name
	ws.cell(row=i+2, column=3).value = '{x:.0f}'.format(x = math.floor(shistory[name]['Score'][CURRENT]))
	ws.cell(row=i+2, column=4).value = '{x:.2f}'.format(x = shistory[name]['Score'][CURRENT]-math.floor(shistory[name]['Score'][CURRENT]))[1:]
	ws.cell(row=i+2, column=6).value = shistory[name]['RM'][CURRENT]
	ws.cell(row=i+2, column=8).value = shistory[name]['RD'][CURRENT]
	ws.cell(row=i+2, column=10).value = shistory[name]['RP'][CURRENT]
	
	if shistory[name]['Score'][CURRENT-1] is not None:
		ws.cell(row=i+2, column=5).value = parity(shistory[name]['Score'][CURRENT]-shistory[name]['Score'][CURRENT-1])
		ws.cell(row=i+2, column=7).value = parity(shistory[name]['RM'][CURRENT]-shistory[name]['RM'][CURRENT-1])
		ws.cell(row=i+2, column=9).value = parity(shistory[name]['RD'][CURRENT]-shistory[name]['RD'][CURRENT-1])
		ws.cell(row=i+2, column=11).value = parity(shistory[name]['RP'][CURRENT]-shistory[name]['RP'][CURRENT-1])
	
	best3 = []
	worst3 = []
	if name in history[CURRENT].keys():
		
		sortedrounds = dict()
		sortedrounds = {k: v for k, v in sorted(history[CURRENT][name].items(),key = lambda item: item[1],reverse = True)}
		sortedroundslist = []
		for rnd, rating in sortedrounds.items():
			sortedroundslist.append([rnd,rating[0],rating[1],rating[2]])
		
		for j in range(min(3,len(sortedroundslist))):
			if sortedroundslist[j][1] > 0:
				best3.append(sortedroundslist[j])
			if sortedroundslist[len(sortedroundslist)-j-1][1]<0:
				worst3.append(sortedroundslist[-j-1])
	
	
	for j in range(0,min(3,len(best3))):
		ws.cell(row = i+2, column = 3*j+13+0).value = best3[j][0]
		ws.cell(row = i+2, column = 3*j+13+1).value = '+' + '{:.1f}'.format(best3[j][1])
		ws.cell(row = i+2, column = 3*j+13+2).value = str(best3[j][2]) + ' / ' + str(best3[j][3])
			
	for j in range(0,min(3,len(worst3))):
		ws.cell(row = i+2, column = 3*j+23+0).value = worst3[j][0]
		ws.cell(row = i+2, column = 3*j+23+1).value = worst3[j][1]
		ws.cell(row = i+2, column = 3*j+23+2).value = str(worst3[j][2]) + ' / ' + str(worst3[j][3])

ws = wb.create_sheet("TWOWs")
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 75
ws['A1'].value = 'Tracked TWOWs'
ws['A1'].style = default
ws['B1'].style = default
ws.row_dimensions[1].height = 23
f = open('data/index.txt','r',encoding = 'utf-8')
d = f.readlines()
f.close()
i = 2
for twow in d:
	ws.row_dimensions[i].height = 23
	ws.cell(row=i, column=1).style = default2
	ws.cell(row=i, column=2).style = default2
	ws.cell(row=i, column=1).value = twow.rstrip()
	i += 1

ws.cell(row=i, column=1).value = 'Banned TWOWs'
ws.cell(row=i, column=2).value = 'Reason'
ws.cell(row=i, column=1).style = default
ws.cell(row=i, column=2).style = default
ws.row_dimensions[i].height = 23
i += 1
f = open('data/blacklist.txt','r',encoding = 'utf-8')
d = f.readlines()
f.close()
j = 1
for line in d:
	ws.row_dimensions[i].height = 23
	ws.cell(row=i, column=j).value = line.rstrip()
	ws.cell(row=i, column=j).style = default2
	j += 1
	if j == 3:
		i += 1
		j = 1


#Rounds
ws = wb.create_sheet("Rounds")

#column sizes
sizes = [35,11]
for i in range(2*CURRENT+2):
	if i < 26:
		ws.column_dimensions[letters[i]].width = sizes[i%2]
	else:
		ws.column_dimensions[letters[i//26-1] + letters[i%26]].width = sizes[i%2]
		
#cell style for whole header
for i in range(2*CURRENT+2):	
	ws.cell(row=1, column=i+1).style = default
	if i%2 == 0:
		val = CURRENT-i//2
		ws.cell(row=1, column=i+1).value = months[val%12]+" \'"+str((val+192)//12)
		
for i in range(CURRENT+1):
	r = 2
	roundlist = []
	for rd in rounds[CURRENT-i]:
		roundlist.append([rd,rounds[CURRENT-i][rd][0]])
	
	roundlist = sorted(roundlist,key=lambda x: (x[1]), reverse = True)
	
	for rd in roundlist:
		ws.cell(row=r, column=2*i+1).style = default2
		ws.cell(row=r, column=2*i+1).value = rd[0]
		
		ws.cell(row=r, column=2*i+2).fill = PatternFill('solid', fgColor=color_bg(rd[1]))
		ws.cell(row=r, column=2*i+2).font = Font(name='Assistant', color=color_text(rd[1]), size=13)
		ws.cell(row=r, column=2*i+2).number_format = '0'
		ws.cell(row=r, column=2*i+2).value = rd[1]
		r += 1
	
	while r < 126:
		ws.cell(row=r, column=2*i+1).style = default2
		ws.cell(row=r, column=2*i+2).style = default2
		r += 1

	



wb.save("glicko.xlsx")
