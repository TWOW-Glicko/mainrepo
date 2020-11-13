import random
import json
import math
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from openpyxl.utils import get_column_letter
from PIL import ImageFont
from datetime import date

months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
days = [31,28,31,30,31,30,31,31,30,31,30,31]
letters = [chr(ord('A')+i) for i in range(26)]

CURRDATE = 200731
CURRYR = 20
CURRMTH = 7
CURRDY = 31

def inttodate(txtdate):
	yr = 2000+txtdate//10000
	mth = (txtdate//100)%100
	dy = txtdate % 100
	rldate = date(yr,mth,dy)
	return rldate


def findrating(data, rdate, name, memo):	
	firstdate = inttodate(data[0])
	diff = rdate - firstdate
	index = diff.days+1
	
	if len(data[index]) == 3:
		rd = data[index][0]
		rm = data[index][1]
		rp = data[index][2]
	elif name in memo.keys():
		rd = data[index][0]
		rm = memo[name][0]
		rp = memo[name][1]
	else:
		rd = data[index][0]
		while len(data[index]) == 1:
			index -= 1
		rm = data[index][1]
		rp = data[index][2]
	
	memo[name] = [rm,rp]
	score = 5*(rm-2*rd)
	
	return (score,rm,rd,rp,memo)
	
with open('resultdaily.json',encoding = 'utf-8') as f: 
    result = json.load(f)			#Dictionary full of contestant stats [Contestant][first date,[raw RM, raw carry RD, RP, raw RD]]

g = open('viddata.json',"w",encoding = 'utf-8')

g.write("{")

memo = dict()

for yr in range(16,CURRYR+1):
	for mth in range(1,13):
		mthdy = days[mth-1]
		if yr%4 == 0 and mth == 2: mthdy += 1
		for dy in range(1,mthdy+1):
			if 10000*yr+100*mth+dy <= CURRDATE:
				print("YEAR " + str(yr) + " MONTH " + str(mth) + " DAY " + str(dy))
				
				if 10000*yr+100*mth+dy != 160101: g.write(",")
				g.write(str(10000*yr+100*mth+dy)+":")
				g.write("{")
				
				ranking = []
				
				for name, ratings in result.items():	#Cycles through all names (rating = [list of [RD, RM, RP]])
					if	ratings[0] <= 10000*yr+100*mth+dy:
						(score,rm,rd,rp,memo) = findrating(ratings,inttodate(10000*yr+100*mth+dy),name,memo)
						temparray = []
						temparray.append(score)
						temparray.append(rm)
						temparray.append(rd)
						temparray.append(rp)
						temparray.append(name)
						ranking.append(temparray)
				
				ranking.sort(reverse=True)						#sort ranking from highest to lowest score (to get actual rankings)
				
				for i in range(50):
					currcnt = ranking[i]
				
					if i != 0: g.write(",")
					g.write(str(i+1)+":")
					g.write("[")
					g.write("\""+ currcnt[4] + "\",")
					g.write(str(round(currcnt[0],3)) + ",")
					g.write(str(round(5*currcnt[1],3)) + ",")
					g.write(str(round(5*currcnt[2],3)) + ",")
					g.write(str(round(currcnt[3],3)) + "]")
					
				g.write("}")
			

g.write("}")

g.close()



