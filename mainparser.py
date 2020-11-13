import json
import os
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('dataparse.xlsx')
ws = wb['ADD TWOW DATA HERE']
ROWS = 1000

if not os.path.isdir('dataparsed'):
	os.makedirs('dataparsed')

for i in range(2, ROWS):
	name = ws.cell(row=1, column=i).value
	rnd = ws.cell(row=2, column=i).value
	
	if name is not None and rnd is not None:
		if not (os.path.isdir('data/'+name) or os.path.isdir('dataparsed/'+name)):
			f = open('data/index.txt', 'a')
			f.write(name+'\n')
			
		if not os.path.isdir('dataparsed/'+name):
			os.makedirs('dataparsed/'+name)	
			
		dy = ws.cell(row=3, column=i).value
		
		print(name, rnd, dy)
			
		f = open('dataparsed/'+name+'/'+str(int(rnd))+'.txt', 'w', encoding = 'utf-8')
		f.write(str(int(dy))+'\n')
		j = 4
		while ws.cell(row=j, column=i).value is not None:
			f.write(ws.cell(row=j, column=i).value+'\n')
			j += 1
		f.close()
