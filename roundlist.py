import random
import json
import math
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from openpyxl.utils import get_column_letter
from PIL import ImageFont

months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']	#What do you think this is
letters = [chr(ord('A')+i) for i in range(26)]

#Give me a month!
print('Month to generate?')
CURRENT = int(input().rstrip())	#CURRENT = month number
print('Start date?')
STDT = int(input().rstrip())
print('End date?')
EDDT = int(input().rstrip())
	
with open('rounds.json',encoding = 'utf-8') as f:
	rounds = json.load(f)		#Dictionary (once nested) full of round strengths [month][Roundname][Strength]

g = open('roundlist.txt','w',encoding = 'utf-8')

mthrds = []


for rd in rounds[CURRENT].keys():
	if STDT <= rounds[CURRENT][rd][1] % 100 <= EDDT:
		mthrds.append([rounds[CURRENT][rd][1],rd])
	
mthrds.sort()

for i in range(len(mthrds)):
	g.write(str(mthrds[i][0]) + "\t" + mthrds[i][1] + "\n")

print(mthrds)

